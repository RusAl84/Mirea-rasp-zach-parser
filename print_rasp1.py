import openpyxl

def str_count(text, substr):    #количество вхождений подстроки в строку
    return len(text.split(substr))-1

theFile = openpyxl.load_workbook('1.xlsx')
#print(theFile.sheetnames)
currentSheet = theFile['Лист1']

 # 49 шифров
 # 17 листов
i=1
while i<300:
    rw=2 # строка с шифрами групп
    gr=str(currentSheet.cell(row=rw, column=i).value)
    #print(gr)
    if len(gr)>4:
         if str_count(gr, '-')>1:
            print(gr)
    i=i+1